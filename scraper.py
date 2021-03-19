import re
import json
import requests
import datetime
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.chart import PieChart, ProjectedPieChart, Reference
from openpyxl.styles import Alignment, Font
from urllib3.exceptions import NewConnectionError, MaxRetryError


# From the json file gets the position and creates a dictionary with techs to search
def get_search_data():
    total_techs = 0
    techs = {}

    try:
        with open('search.json') as file:
            data = json.load(file)

        position = data.get('position')
        techs_list = data.get('techs')
        for tech in techs_list:
            techs.update({tech:0})
            total_techs += 1

        file.close()

    except FileNotFoundError:
        print('ERROR: search.json not found.')

    return position,techs,total_techs

# Generates the url from position
def get_url(position):
    template = 'https://www.indeed.com/jobs?q={}&start=00'
    url = template.format(position)

    return url

# Generates a string with the desciption of the job post
def get_job_description(post):

    try:
        job_url = 'https://www.indeed.com' + post.h2.a.get('href')
        response = requests.get(job_url)
        soup = BeautifulSoup(response.text, 'html.parser')

        try:
            job_desc = soup.find('div', 'jobsearch-jobDescriptionText').text
        except AttributeError:
            print("job_desc NONE") #debugger
            job_desc = None

    except (ConnectionError, TimeoutError, NewConnectionError, MaxRetryError):
        print("\n\nERROR: Connection failed trying to get the description of a job.\n\n") #debugger
        job_desc = None
    
    return job_desc

# Counts the techs that are in the job description and updates the techs dictionary
def count_techs(description,techs):
    total_found = 0
    techs_to_increment = []
    job_desc = re.split(r'[-,/().\s]\s*',description)

    #for each word in the description, if it is equals to a tech, that tech has to be incremented
    for word in job_desc:
        word_lower = word.lower()
        for tech in techs:
            if word_lower == tech:
                if tech not in techs_to_increment:
                    techs_to_increment.append(tech)

    for tech in techs_to_increment:
        techs[tech] += 1
        total_found += 1
    
    return total_found

# Modifies the dictionary joining the words that refers to the same techonoly/language
def join_techs(techs_dict):

    try:
        with open('search.json') as file:
            data = json.load(file)

        try:
            to_join_list = data.get('join')
            for raw_techs_to_join in to_join_list:
                techs_to_join = raw_techs_to_join.split(":")

                #gets the keys and values
                tech_to_be_updated = techs_to_join[0]
                value_to_be_updated = techs_dict.get(tech_to_be_updated)

                tech_to_del = techs_to_join[1]
                value_to_save = techs_dict.get(tech_to_del)

                new_value = value_to_be_updated + value_to_save
                
                #modifies the dictionary
                techs_dict.update({tech_to_be_updated:new_value})
                techs_dict.pop(tech_to_del)
                
        except (TypeError,IndexError):
                print("ERROR: 'join' config.")

        file.close()
    
    except FileNotFoundError:
        print('ERROR: search.json not found.')

# Sort the techs dictionary descendingly
def sort_techs(techs_dict):
    sorted_techs_dict = {}
    sorted_keys = sorted(techs_dict, key=techs_dict.get, reverse=True)

    for key in sorted_keys:
        #sorted_techs_dict.update({key:techs_dict.get(key)})
        sorted_techs_dict[key] = techs_dict[key]
    
    return sorted_techs_dict
    
# Generates a .xlsx file with the results of the scraping
def results_to_excel(position,techs,posts_seen,total_found):
    wb = Workbook()
    sheet = wb.active
    date = datetime.datetime.now()

    # initial format of the sheet
    sheet.cell(1,1).value = 'Total posts viewed:'
    sheet.cell(1,2).value = posts_seen
    sheet.cell(1,4).value = 'Date:'
    sheet.cell(1,5).value = date
    sheet.cell(3,1).value = 'Tech/Lang'
    sheet.cell(3,2).value = 'Number of appearances'
    sheet.cell(3,3).value = 'Ocurrence percentage (*)'
    sheet['F19'] = "(*) respect to the total number of posts viewed"
    sheet.cell(1,1).font = Font(name='Arial', bold=True, size=13)
    sheet.cell(1,4).font = Font(name='Arial', bold=True, size=13)
    sheet.cell(3,1).font = Font(name='Arial', bold=True, size=13)
    sheet.cell(3,2).font = Font(name='Arial', bold=True, size=13)
    sheet.cell(3,3).font = Font(name='Arial', bold=True, size=13)
    sheet['F19'].font = Font(name='Arial', size=13)

    # dump the techs dictionary the sheet
    row = 4
    for key,value in techs.items():
        sheet.cell(row,1).value = key
        sheet.cell(row,2).value = value
        sheet.cell(row,3).value = f'{"{:.2f}".format(value*100/posts_seen)}%' 
        sheet.cell(row,3).alignment = Alignment(horizontal='right')
        row += 1

    # generates the chart of the results
    pie = PieChart()
    labels = Reference(sheet, min_col=1, min_row=4, max_row=row-1)
    data = Reference(sheet, min_col=2, min_row=4, max_row=row-1)
    pie.add_data(data)
    pie.set_categories(labels)
    pie.title = f'Number of times a technology/language was found from {total_found} techs found'

    sheet.add_chart(pie, 'F3')

    wb.save(f'{position} {date}.xlsx')

def main():
    posts_seen = 0
    total_found = 0 #total number of times that all technolgies/languages where found
    position,techs,total_techs = get_search_data()
    url = get_url(position)

    '''response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    posts = soup.find_all('div', 'jobsearch-SerpJobCard')
    
    i=1 #debugger
    for post in posts:
        posts_seen += 1
        job_desc = get_job_description(post)
        if job_desc != None:
            total_found += count_techs(job_desc,techs)
        print(f"{i}: {techs}") #debugger
        i += 1 #debugger'''
    
    while True:

        try:
            response = requests.get(url)
            soup = BeautifulSoup(response.text, 'html.parser')
            posts = soup.find_all('div', 'jobsearch-SerpJobCard')
            i=1 #debugger
            for post in posts:
                posts_seen += 1
                job_desc = get_job_description(post)
                if job_desc != None:
                    total_found += count_techs(job_desc,techs)
                print(f"{i}: {techs}") #debugger
                i += 1 #debugger
            
            #Next page
            try:
                url = 'https://www.indeed.com' + soup.find('a', {'aria-label': 'Next'}).get('href')
                print(url) #debugger
            except AttributeError:
                print("\nNext page not found.") #debugger
                break
        
        except (ConnectionError, TimeoutError, NewConnectionError, MaxRetryError):
            print("\n\nERROR: Connection failed trying to get the actual page.\n\n") #debugger
            job_desc = None
    
    join_techs(techs)
    #print(techs) #debugger
    sorted_techs = sort_techs(techs)
    #print(sorted_techs) #debbuger
    results_to_excel(position,sorted_techs,posts_seen,total_found)

    print('\nScraping finished, an Excel sheet with the results has been created.\n')


if __name__ == '__main__':
    main()
